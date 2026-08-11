import json
import os
import ezdxf
import tempfile
import traceback
import google.generativeai as genai
from dotenv import load_dotenv

# .env 파일 로드
load_dotenv()
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from firebase_functions import https_fn
from firebase_admin import initialize_app

initialize_app()

def set_cors_headers(resp: https_fn.Response):
    """CORS 헤더를 항상 추가하는 헬퍼 함수"""
    resp.headers["Access-Control-Allow-Origin"] = "*"
    resp.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
    return resp

@https_fn.on_request(max_instances=10, timeout_sec=120, memory=1024)
def analyze_floorplan_3d(req: https_fn.Request) -> https_fn.Response:
    """프론트엔드에서 2D 3D 변환기의 Gemini 호출을 중계 (API Key 보호)"""
    if req.method == "OPTIONS":
        return set_cors_headers(https_fn.Response(status=204))
    
    if req.method != "POST":
        resp = https_fn.Response(json.dumps({"detail": "Only POST method is supported"}), status=405)
        return set_cors_headers(resp)
        
    try:
        data = req.get_json(silent=True)
        if not data or "imageBase64" not in data or "mimeType" not in data:
            return set_cors_headers(https_fn.Response(json.dumps({"detail": "Missing imageBase64 or mimeType"}), status=400, content_type="application/json"))
            
        api_key = os.environ.get("GOOGLE_API_KEY_EDITOR")
        if not api_key:
             print("[ERROR] GOOGLE_API_KEY_EDITOR missing")
             return set_cors_headers(https_fn.Response(json.dumps({"detail": "Server API Key Missing (Editor)"}), status=400, content_type="application/json"))
             
        # 디버깅: 가용 모델 리스트 출력 (한 번만 혹은 주기적으로)
        import datetime
        print(f"[DEBUG] Starting analyze_floorplan_3d at {datetime.datetime.now()}")
        
        genai.configure(api_key=api_key)
        
        # 모델 리스트 확인 (로그에서 확인용)
        try:
            available_models = [m.name for m in genai.list_models() if 'generateContent' in m.supported_generation_methods]
            print(f"[DEBUG] Available models: {available_models}")
        except Exception as me:
            print(f"[DEBUG] Failed to list models: {str(me)}")
        
        
        system_instruction = """
          너는 전문 건축 CAD AI다. 주어진 평면도 이미지를 분석하여 
          벽(walls), 문(doors), 창문(windows)의 상대적 좌표(x, y)와 크기를 극도로 정확하게 추출하라.
          
          원점(0, 0)은 도면의 좌측 상단 또는 좌측 하단을 기준으로 일관되게 적용하라.
          좌표 단위는 밀리미터(mm)를 가정하고 논리적인 픽셀 비율로 환산해서 제공하라.
          
          반드시 아래의 JSON 스키마 구조와 100% 일치하게 반환해야 하며, 마크다운 코드 블록(```json) 등 그 어떤 부가적인 텍스트도 포함하지 마라.
          
          [JSON 스키마]
          {
            "walls": [
              {
                "id": "고유문자열 (예: wall-1)",
                "start": { "x": 숫자, "y": 숫자 },
                "end": { "x": 숫자, "y": 숫자 },
                "thickness": 숫자 (벽 두께),
                "height": 숫자 (벽 높이, 기본값 2800)
              }
            ],
            "doors": [
               // 생략 (프론트엔드와 동일)
            ],
            "windows": [
               // 생략
            ]
          }
        """
        
        # 모델 설정 (JSON 반환 설정 파이썬 SDK 방식)
        model = genai.GenerativeModel(
            model_name='gemini-2.5-flash',
            generation_config=genai.types.GenerationConfig(
                response_mime_type="application/json",
                temperature=0.1,
            ),
            system_instruction=system_instruction
        )
        
        # Base64 처리 모듈(안전하게 bytes로 디코딩)
        import base64
        base64_data = data["imageBase64"].replace("data:image/jpeg;base64,", "").replace("data:image/png;base64,", "")
        image_bytes = base64.b64decode(base64_data)
        
        prompt = "첨부된 평면도 이미지를 분석하여 지정된 JSON 포맷으로 구조화된 공간 데이터를 추출해 줘."
        
        # 파이썬 SDK에서는 MIME dict(Blob) 형태로 전송
        image_blob = {
            "mime_type": data["mimeType"],
            "data": image_bytes
        }
        
        print(f"[DEBUG] Calling Gemini API with model: {model.model_name}")
        response = model.generate_content([prompt, image_blob])
        print("[DEBUG] Gemini API call successful")
        ai_response_text = response.text.strip()
        
        # 불필요한 마크다운 제거
        clean_json_text = ai_response_text.replace("```json", "").replace("```", "").strip()
        
        # 파싱 가능한 JSON인지 체크 후 리턴 (실패하더라도 일단 클라이언트로 전달 후 클라이언트에서 에러 처리)
        resp = https_fn.Response(json.dumps({"success": True, "data": json.loads(clean_json_text)}), content_type="application/json")
        return set_cors_headers(resp)
        
    except Exception as e:
        error_msg = traceback.format_exc()
        print(f"Gemini API 3D Converter Error: {error_msg}")
        resp = https_fn.Response(json.dumps({"detail": f"AI Parsing Failed: {str(e)}"}), status=400, content_type="application/json")
        return set_cors_headers(resp)

@https_fn.on_request(max_instances=10, memory=1024)
def analyze_dxf(req: https_fn.Request) -> https_fn.Response:
    """DXF 도면 파일 파싱 및 수량(길이, 면적) 추출 클라우드 함수"""
    if req.method == "OPTIONS":
        return set_cors_headers(https_fn.Response(status=204))
    
    if req.method != "POST":
        resp = https_fn.Response(json.dumps({"detail": "Only POST method is supported"}), status=405)
        return set_cors_headers(resp)
        
    try:
        # multipart/form-data 파싱
        # (단순화를 위해 첫 번째 파일 데이터만 추출)
        files = req.files
        if not files or "file" not in files:
             resp = https_fn.Response(json.dumps({"detail": "No file uploaded"}), status=400)
             return set_cors_headers(resp)

        file_obj = files["file"]
        content = file_obj.read()
        filename = file_obj.filename

        with tempfile.NamedTemporaryFile(delete=False, suffix=".dxf") as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        
        try:
            doc = ezdxf.readfile(tmp_path)
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

        msp = doc.modelspace()
        results = {}
        
        for entity in msp:
            layer = entity.dxf.layer
            if layer not in results: results[layer] = {"length": 0.0, "area": 0.0, "count": 0}
            
            etype = entity.dxftype()
            if etype == 'LINE':
                start, end = entity.dxf.start, entity.dxf.end
                results[layer]["length"] += ((end[0]-start[0])**2 + (end[1]-start[1])**2)**0.5
            elif etype == 'LWPOLYLINE':
                points = entity.get_points()
                for i in range(len(points) - 1):
                    p1, p2 = points[i], points[i+1]
                    results[layer]["length"] += ((p2[0]-p1[0])**2 + (p2[1]-p1[1])**2)**0.5
                if entity.is_closed:
                    p1, p2 = points[-1], points[0]
                    results[layer]["length"] += ((p2[0]-p1[0])**2 + (p2[1]-p1[1])**2)**0.5
                    try: results[layer]["area"] += entity.area()
                    except: pass
            results[layer]["count"] += 1
            
        for l in results:
            results[l]["length"] = round(results[l]["length"], 3)
            results[l]["area"] = round(results[l]["area"], 3)
            
        resp = https_fn.Response(json.dumps({"filename": filename, "layers": results}), content_type="application/json")
        return set_cors_headers(resp)

    except Exception as e:
        error_msg = traceback.format_exc()
        print(f"Error Analyzing DXF: {error_msg}")
        resp = https_fn.Response(json.dumps({"detail": f"DXF Parsing Error: {str(e)}"}), status=400, content_type="application/json")
        return set_cors_headers(resp)

@https_fn.on_request()
def recommend_basis(req: https_fn.Request) -> https_fn.Response:
    if req.method == "OPTIONS":
        return set_cors_headers(https_fn.Response(status=204))
        
    item_name = req.args.get('item_name', '')
    if not item_name:
         return set_cors_headers(https_fn.Response(json.dumps({"detail": "Item name is required"}), status=400, content_type="application/json"))

    try:
        api_key = os.environ.get("GOOGLE_API_KEY_SAFETY")
        if not api_key:
             raise ValueError("GOOGLE_API_KEY_SAFETY environment variable is missing")
             
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.5-flash')
        
        prompt = f"당신은 건설 부대토목 전문가입니다. '{item_name}' 작업에 가장 적합한 대한민국 건설공사 표준품셈 (또는 전문 시방서) 기준 조항을 추천해주십시오. 설명 없이 관련된 기준명이나 조항(예: 토목공사 표준품셈 [2-1-1])만을 매우 간결하게 핵심만 대답하세요."
        response = model.generate_content(prompt)
        ai_recommendation = response.text.strip()
        
        resp = https_fn.Response(json.dumps({"item": item_name, "recommended_basis": ai_recommendation}), content_type="application/json")
        return set_cors_headers(resp)
        
    except Exception as e:
        error_msg = traceback.format_exc()
        print(f"Gemini API Error: {error_msg}")
        resp = https_fn.Response(json.dumps({"detail": f"AI Recommendation Failed: {str(e)}"}), status=400, content_type="application/json")
        return set_cors_headers(resp)

@https_fn.on_request(max_instances=10)
def export_excel(req: https_fn.Request) -> https_fn.Response:
    if req.method == "OPTIONS":
        return set_cors_headers(https_fn.Response(status=204))
    
    if req.method != "POST":
         return set_cors_headers(https_fn.Response("Method Not Allowed", status=405))
         
    try:
        data = req.get_json(silent=True) or {}
        wb = Workbook()
        ws = wb.active
        ws.title = "수량산출서"
        
        header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        headers = ["No", "공종/항목", "규격", "단위", "수량", "산출근거"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            
        items = data.get("items", [])
        for i, item in enumerate(items, start=1):
            row = [i, item.get("name"), item.get("spec"), item.get("unit"), item.get("quantity"), item.get("basis")]
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if isinstance(cell.value, (int, float)) else "left")

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(tmp.name)
        tmp_path = tmp.name
        tmp.close()
        
        with open(tmp_path, "rb") as f:
            file_data = f.read()
            
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
            
        resp = https_fn.Response(
            file_data, 
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=Namhwa_Quantity_Report.xlsx"}
        )
        return set_cors_headers(resp)
    except Exception as e:
        error_msg = traceback.format_exc()
        print(f"Excel Export Error: {error_msg}")
        resp = https_fn.Response(json.dumps({"detail": f"Excel Export Error: {str(e)}"}), status=500)
        return set_cors_headers(resp)

@https_fn.on_request(max_instances=10, timeout_sec=60, memory=512)
def generate_shm_summary(req: https_fn.Request) -> https_fn.Response:
    """SHM System 체크리스트 점수 및 코멘트를 바탕으로 Gemini AI 맞춤형 종합 요약 생성"""
    if req.method == "OPTIONS":
        return set_cors_headers(https_fn.Response(status=204))
    if req.method != "POST":
        return set_cors_headers(https_fn.Response(json.dumps({"detail": "Only POST method is supported"}), status=405, content_type="application/json"))
        
    try:
        data = req.get_json(silent=True) or {}
        site_name = data.get("siteName", "현장명 미상")
        final_score = data.get("finalScore", 0)
        strengths = data.get("strengths", [])
        weaknesses = data.get("weaknesses", [])
        
        api_key = os.environ.get("GOOGLE_API_KEY_SAFETY")
        if not api_key:
             raise ValueError("GOOGLE_API_KEY_SAFETY environment variable is missing")
             
        genai.configure(api_key=api_key)
        
        system_instruction = """
        당신은 20년 경력의 심도 있는 지식을 갖춘 대한민국 최고 건설현장 안전보건 전문 심사위원(책임기술인)입니다.
        엄격하지만 긍정적이고 따뜻한 카리스마를 가졌습니다.
        현장의 안전보건 시스템 점검 결과를 분석하여, 건설현장 책임자들에게 제공될 '총평 및 개선 요구사항 종합 보고서'를 작성하세요.
        
        [작성 규칙]
        1. 출력 형식은 순수 HTML 포맷(<b>, <i>, <br>, <span style="..."> 등)만 사용해야 하며, 마크다운 코드 블록(```html) 등 부수적인 텍스트는 응답에 포함하지 마십시오.
        2. 첫 문장은 제공된 점수를 바탕으로 현재 현장의 안전보건 활동 적정성 수준(우수/양호/미흡 등)에 대해 한 줄로 강력하게 총평합니다.
        3. 강점 분야가 있다면 어떻게 현장을 긍정적으로 이끌고 있는지 칭찬해주십시오.
        4. 약점 분야(사용자 코멘트 포함)가 있다면 사고 예방을 위해 구체적이고 전문적으로 어떻게 개선해야 할지 따끔한 조언을 남기세요.
        5. 마지막은 구성원들의 안전 의식을 끓어오르게 할 독창적이고 힘찬 형태의 짧은 안전 슬로건으로 마무리하세요.
        6. 전체 텍스트 양은 A4 용지 4분의 1장이 넘지 않게 간결하고 가독성 좋게, 줄바꿈을 적절히 사용하여 구성하세요.
        7. 똑같은 점수와 비슷한 내용이 들어오더라도 매번 단어와 슬로건을 다르게 구성하여 중복 느낌을 확실히 피하십시오.
        """
        
        model = genai.GenerativeModel(
            model_name='gemini-2.5-flash',
            generation_config=genai.types.GenerationConfig(
                temperature=0.8, # 다양성 창출을 위해 온도 살짝 상향
            ),
            system_instruction=system_instruction
        )
        
        # 프롬프트 조립
        weaknesses_str = json.dumps(weaknesses, ensure_ascii=False) if weaknesses else "특별한 취약점 코멘트 없음"
        strengths_str = ", ".join(strengths) if strengths else "강점 분야로 꼽을 만한 사항 미흡"
        
        prompt = f"현장명: {site_name}\n총점: {final_score}점\n강점섹션: {strengths_str}\n약점섹션 및 세부 지적사항: {weaknesses_str}\n\n위 데이터를 바탕으로 종합 분석 HTML 텍스트를 응답해라. <b>[안전보건 활동 적정성 종합 검토]</b> 라는 제목으로 시작해라."
        
        response = model.generate_content(prompt)
        ai_summary = response.text.strip()
        
        # 앞뒤 마크다운 트림
        ai_summary = ai_summary.replace("```html", "").replace("```", "").strip()
        
        resp = https_fn.Response(json.dumps({"success": True, "summary": ai_summary}), content_type="application/json")
        return set_cors_headers(resp)
        
    except Exception as e:
        error_msg = traceback.format_exc()
        print(f"Gemini SHM Summary API Error: {error_msg}")
        resp = https_fn.Response(json.dumps({"success": False, "detail": f"AI 분석 중 오류 발생: {str(e)}"}), status=400, content_type="application/json")
        return set_cors_headers(resp)