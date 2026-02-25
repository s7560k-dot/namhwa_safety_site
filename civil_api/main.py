from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
import uvicorn
import os
import ezdxf
import io
import tempfile
import pandas as pd
from typing import Optional, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

app = FastAPI(title="Namhwa Civil Engineering API")

# CORS 설정 (프론트엔드 통신 허용)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
async def root():
    return {"message": "Welcome to Namhwa Civil Engineering Quantity Takeoff API"}

# --- [Phase 1] 기본 산출 로직 ---
@app.get("/calculate/earthwork")
async def calculate_earthwork(length: float, width: float, depth: float, slope_ratio: float = 0.3):
    """관로 터파기 수량 산출"""
    try:
        top_width = width + (2 * depth * slope_ratio)
        area = ((width + top_width) / 2) * depth
        total_volume = area * length
        basis = f"(({width} + {top_width}) / 2) * {depth} * {length}"
        
        return {
            "item": "관로 터파기 (토사)",
            "quantity": round(total_volume, 2),
            "unit": "m3",
            "basis": basis,
            "parameters": {"length": length, "bottom_width": width, "top_width": top_width, "depth": depth, "slope_ratio": slope_ratio}
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# --- [Phase 2] DXF/DWG 도면 분석 ---
@app.post("/analyze/dxf")
async def analyze_dxf(file: UploadFile = File(...), layer_filter: Optional[str] = None):
    """DXF/DWG 도면에서 선 길이 및 면적 추출"""
    temp_dxf_path = None
    try:
        filename = file.filename.lower()
        content = await file.read()
        
        # DWG 처리 (DXF로 변환)
        if filename.endswith(".dwg"):
            try:
                import aspose.cad as cad
                from aspose.cad import Image
                from aspose.cad.imageoptions import DxfOptions
                
                with tempfile.NamedTemporaryFile(delete=False, suffix=".dwg") as tmp_dwg:
                    tmp_dwg.write(content)
                    tmp_dwg_path = tmp_dwg.name
                
                temp_dxf_path = tmp_dwg_path.replace(".dwg", "_converted.dxf")
                image = Image.load(tmp_dwg_path)
                image.save(temp_dxf_path, DxfOptions())
                
                doc = ezdxf.readfile(temp_dxf_path)
                os.remove(tmp_dwg_path)
            except ImportError:
                raise HTTPException(status_code=400, detail="DWG 변환을 위해 aspose-cad 라이브러리가 필요합니다. (Python 3.12 이하 권장)")
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"DWG 변환 실패: {str(e)}")
        else:
            # DXF 직접 처리
            doc = ezdxf.read(io.BytesIO(content))

        msp = doc.modelspace()
        results = {}
        
        for entity in msp:
            layer = entity.dxf.layer
            if layer_filter and layer_filter not in layer: continue
            if layer not in results: results[layer] = {"length": 0.0, "area": 0.0, "count": 0}
            
            etype = entity.dxftype()
            if etype == 'LINE':
                start, end = entity.dxf.start, entity.dxf.end
                results[layer]["length"] += ((end[0]-start[0])**2 + (end[1]-start[1])**2)**0.5
            elif etype == 'LWPOLYLINE':
                points = entity.get_points()
                for i in range(len(points) - 1):
                    p1, p2 = points[i], points[i+1]
                    results[layer]["length"] += ((p2[0]-p1[0])**2 + (p2[1]-p1[1])**2)**0.5
                if entity.is_closed:
                    p1, p2 = points[-1], points[0]
                    results[layer]["length"] += ((p2[0]-p1[0])**2 + (p2[1]-p1[1])**2)**0.5
                    try: results[layer]["area"] += entity.area()
                    except: pass
            results[layer]["count"] += 1
            
        for l in results:
            results[l]["length"] = round(results[l]["length"], 3)
            results[l]["area"] = round(results[l]["area"], 3)
            
        return {"filename": file.filename, "layers": results}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"도면 분석 에러: {str(e)}")
    finally:
        if temp_dxf_path and os.path.exists(temp_dxf_path):
            os.remove(temp_dxf_path)

# --- [Phase 3] AI 기반 근거 추천 (RAG 준비) ---
@app.get("/ai/recommend-basis")
async def recommend_basis(item_name: str):
    """표준품셈 기반 산출 근거 추천 (Mockup)"""
    recommendations = {
        "터파기": "토목공사 표준품셈 [2-1-1] 인력터파기 및 되메우기 기준 적용",
        "포장": "도로공사 표준시방서 제5장 아스팔트 콘크리트 포장공사 기준 적용",
        "측구": "토목공사 표준품셈 [4-2-1] 콘크리트 측구 설치 기준 적용"
    }
    basis = recommendations.get(item_name, "관련 품셈 근거 검색 중...")
    return {"item": item_name, "recommended_basis": basis}

# --- [Phase 4] 엑셀 보고서 생성 ---
@app.post("/export/excel")
async def export_excel(data: dict):
    """산출 데이터를 엑셀 파일로 변환"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "수량산출서"
        
        # 스타일 정의
        header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        headers = ["No", "공종/항목", "규격", "단위", "수량", "산출근거"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            
        items = data.get("items", [])
        for i, item in enumerate(items, start=1):
            row = [i, item.get("name"), item.get("spec"), item.get("unit"), item.get("quantity"), item.get("basis")]
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if isinstance(cell.value, (int, float)) else "left")

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(tmp.name)
        tmp_path = tmp.name
        tmp.close()
        
        return FileResponse(tmp_path, filename="Namhwa_Quantity_Report.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel Export Error: {str(e)}")

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
